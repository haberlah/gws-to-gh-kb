#!/usr/bin/env python3
"""Convert local .docx → .md (with images) and .xlsx → .csv per tab.

Processes all .docx and .xlsx files in a directory (recursively).
- .docx → .md via pandoc, images extracted to images/ subfolder
- .xlsx → subdirectory with .csv per sheet tab

Usage: python3 convert_local.py <directory>

Requirements: pandoc on PATH, openpyxl (pip install openpyxl)
"""
import subprocess
import sys
import os
import csv

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip3 install openpyxl")
    sys.exit(1)


def convert_docx(docx_path):
    """Convert a .docx file to .md with images extracted."""
    directory = os.path.dirname(os.path.abspath(docx_path))
    basename = os.path.splitext(os.path.basename(docx_path))[0]
    md_path = os.path.join(directory, f"{basename}.md")
    images_dir = os.path.join(directory, "images")

    # Check if pandoc is available
    try:
        subprocess.run(["pandoc", "--version"], capture_output=True, check=True)
    except (FileNotFoundError, subprocess.CalledProcessError):
        print(f"  SKIP: pandoc not found")
        return False

    # Convert with pandoc, extracting media to images/
    os.makedirs(images_dir, exist_ok=True)
    result = subprocess.run(
        ["pandoc", docx_path, "-t", "markdown", "--wrap=none",
         f"--extract-media={images_dir}"],
        capture_output=True, text=True
    )

    if result.returncode != 0:
        print(f"  FAIL: {result.stderr.strip()}")
        return False

    # Write markdown
    with open(md_path, "w") as f:
        f.write(result.stdout)

    sanitize_script = os.path.join(os.path.dirname(__file__), "sanitize_markdown.py")
    if os.path.exists(sanitize_script):
        subprocess.run([sys.executable, sanitize_script, md_path], check=False)

    # Check if any images were extracted
    media_dir = os.path.join(images_dir, "media")
    img_count = 0
    if os.path.exists(media_dir):
        img_count = len([f for f in os.listdir(media_dir) if not f.startswith(".")])

    # Clean up empty images dir if no images
    if img_count == 0:
        try:
            os.rmdir(images_dir)
        except OSError:
            pass

    size = os.path.getsize(md_path)
    print(f"  → {basename}.md ({size:,} bytes, {img_count} images)")
    return True


def convert_xlsx(xlsx_path):
    """Convert an .xlsx file to .csv per sheet tab."""
    directory = os.path.dirname(os.path.abspath(xlsx_path))
    basename = os.path.splitext(os.path.basename(xlsx_path))[0]
    sheet_dir = os.path.join(directory, basename)
    os.makedirs(sheet_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name.replace("/", "_").replace(":", "_")
        csv_path = os.path.join(sheet_dir, f"{safe_name}.csv")

        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        size = os.path.getsize(csv_path)
        print(f"  → {basename}/{safe_name}.csv ({size:,} bytes)")

    wb.close()
    return True


def process_directory(directory):
    """Process all .docx and .xlsx files in a directory recursively."""
    docx_count = 0
    xlsx_count = 0

    for root, dirs, files in os.walk(directory):
        for fname in sorted(files):
            fpath = os.path.join(root, fname)
            rel = os.path.relpath(fpath, directory)

            if fname.lower().endswith(".docx"):
                # Skip if .md already exists with same basename
                md_path = os.path.join(root, os.path.splitext(fname)[0] + ".md")
                if os.path.exists(md_path):
                    print(f"SKIP (md exists): {rel}")
                    continue
                print(f"DOCX → MD: {rel}")
                if convert_docx(fpath):
                    docx_count += 1

            elif fname.lower().endswith(".xlsx"):
                # Skip if csv subdirectory already exists
                csv_dir = os.path.join(root, os.path.splitext(fname)[0])
                if os.path.exists(csv_dir) and any(f.endswith(".csv") for f in os.listdir(csv_dir)):
                    print(f"SKIP (csvs exist): {rel}")
                    continue
                print(f"XLSX → CSV: {rel}")
                if convert_xlsx(fpath):
                    xlsx_count += 1

    print(f"\nConverted {docx_count} .docx and {xlsx_count} .xlsx files")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 convert_local.py <directory>")
        sys.exit(1)
    process_directory(sys.argv[1])
